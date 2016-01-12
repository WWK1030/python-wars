# coding=utf-8

import requests
from bs4 import BeautifulSoup
import re
from  openpyxl.workbook  import  Workbook
from  openpyxl.writer.excel  import  ExcelWriter
from  openpyxl.cell  import  get_column_letter

url = "http://movie.douban.com/top250"


class DouBanMovies:
    def __init__(self):
        self.movies_rank = ''
        self.movies_name = ''
        self.each_info_director = ''
        self.each_info_date = ''
        self.movies_director = ''
        self.movies_date = ''
        self.movies_rate = ''
        self.movies_country = ''
        self.movies_classification = ''

    def movies_info(self, start):
        params = {'start': start, 'filter': ''}
        result = requests.get(url, params=params)
        return result.text

    def movies(self):
        wb = Workbook()
        dest_filename = 'Douban_top250.xlsx'
        ws = wb.active
        ws.title = "Douban_250"
        ws.cell(row=1, column=1, value='Rank')
        ws.cell(row=1, column=2, value='Name')
        ws.cell(row=1, column=3, value='Director')
        ws.cell(row=1, column=4, value='Date')
        ws.cell(row=1, column=5, value='Rate')
        ws.cell(row=1, column=6, value='Country')
        ws.cell(row=1, column=7, value='Classification')
        for start in range(0, 250, 25):
            soup = BeautifulSoup(self.movies_info(str(start)), "html.parser")
            all_info = soup.ol.find_all('li')
            for each_item in all_info:
                self.movies_rank = each_item.find('em').contents[0]
                self.movies_name = each_item.find('span', attrs={'class': 'title'}).contents[0].strip()
                self.each_info_director = each_item.p.contents[0]
                # print(self.each_movies_info)
                director_info = re.findall('导演: (.*)', self.each_info_director)
                for self.movies_director in director_info:
                    if "主" in self.movies_director:
                        self.movies_director = re.findall('导演: (.*)主', self.each_info_director)[0].strip()
                    else:
                        self.movies_director = director_info[0].strip()
                self.each_info_date = each_item.p.contents[1].text.replace('\n', '').strip()
                # print(self.each_info_date)
                movies_info = self.each_info_date.split('/')
                # print(movies_info)
                self.movies_date = '/'.join(movies_info[0:-2])
                self.movies_rate = each_item.find('span', attrs={'class': 'rating_num'}).contents[0]
                self.movies_country = movies_info[-2].strip()
                self.movies_classification = movies_info[-1].strip()
                print(self.movies_rank, self.movies_name, self.movies_director, self.movies_date, self.movies_rate,
                      self.movies_country, self.movies_classification)
                ws.cell(row=int(self.movies_rank) + 1, column=1, value=self.movies_rank)
                ws.cell(row=int(self.movies_rank) + 1, column=2, value=self.movies_name)
                ws.cell(row=int(self.movies_rank) + 1, column=3, value=self.movies_director)
                ws.cell(row=int(self.movies_rank) + 1, column=4, value=self.movies_date)
                ws.cell(row=int(self.movies_rank) + 1, column=5, value=self.movies_rate)
                ws.cell(row=int(self.movies_rank) + 1, column=6, value=self.movies_country)
                ws.cell(row=int(self.movies_rank) + 1, column=7, value=self.movies_classification)
        wb.save(filename = dest_filename)


excel_data = DouBanMovies()
excel_data.movies()
