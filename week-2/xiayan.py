# -*- coding:utf-8 -*-
import requests
from bs4 import BeautifulSoup
import xlwt
from openpyxl import Workbook

URL = 'http://movie.douban.com/top250'


def get_a_page(start):
    param = {'start': start}
    return requests.get(URL, param).content


def get_every_info(data):
    every_info = []
    soup = BeautifulSoup(data)
    lis = soup.find('ol').find_all('li')
    for li in lis:
        name = li.find('span', attrs={'class': 'title'}).text
        info = li.find('div', attrs={'class': 'bd'}).text

        infos = [x.strip() for x in info.strip().split('\n') if len(x) > 0]
        infos = [x for x in infos if len(x) > 0]
        year = infos[1].split('/')[0]
        director = infos[0].split('导演:')[1].split('主演')[0]
        region = infos[1].split('/')[1]
        category = infos[1].split('/')[2]
        star = li.find('span', attrs={'class': 'rating_num'}).text
        print("name: %s, director %s, region %s, year %s, star %s, category %s" % (
            name, director, region, year, star, category))
        every_info.append((name, director, region, year, star, category))
    return every_info


if __name__ == '__main__':

    def write_to_excel(i, name, director, region, year, star, category):
        i = str(i + 2)
        ws['A' + i] = str(i)
        ws['B' + i] = str(name)
        ws['C' + i] = str(director)
        ws['D' + i] = str(region)
        ws['E' + i] = str(year)
        ws['F' + i] = str(star)
        ws['G' + i] = str(category)


    wb = Workbook()
    ws = wb.active

    ws['A1'] = '序号'
    ws['B1'] = '电影名称'
    ws['C1'] = '导演'
    ws['D1'] = '地区'
    ws['E1'] = '上映时间'
    ws['F1'] = '评分'
    ws['G1'] = '类型'

    infos = []
    for i in range(0, 250, 25):
        infos.extend(get_every_info(get_a_page(i)))
    for i in range(0, len(infos)):
        write_to_excel(i, infos[i][0], infos[i][1], infos[i][2], infos[i][3], infos[i][4], infos[i][5])
    wb.save('movie.xlsx')
